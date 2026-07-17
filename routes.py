# =============================================
# WardPulse AI - All Routes (Blueprints)
# =============================================
# Organized into: auth, main, user, admin, api
# Each blueprint handles a specific domain

import os
import io
import json
from datetime import datetime
from functools import wraps
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, current_app, send_from_directory, send_file, session
)
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from models import db, User, Admin, Issue, Vote, Notification


# =============================================
# Blueprint Definitions
# =============================================
auth_bp = Blueprint('auth', __name__)
main_bp = Blueprint('main', __name__)
user_bp = Blueprint('user', __name__)
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')
api_bp = Blueprint('api', __name__, url_prefix='/api')


# =============================================
# Helper Functions
# =============================================
def allowed_file(filename):
    """Check if uploaded file has an allowed extension"""
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']


def admin_required(f):
    """Decorator to require admin login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_id'):
            flash('Admin access required.', 'error')
            return redirect(url_for('auth.admin_login'))
        admin = Admin.query.get(session['admin_id'])
        if not admin or not admin.is_active_admin:
            session.pop('admin_id', None)
            flash('Admin access denied.', 'error')
            return redirect(url_for('auth.admin_login'))
        return f(*args, **kwargs)
    return decorated_function


def super_admin_required(f):
    """Decorator to require super admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_id'):
            return redirect(url_for('auth.admin_login'))
        admin = Admin.query.get(session['admin_id'])
        if not admin or not admin.is_super_admin:
            flash('Super Admin access required.', 'error')
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def get_current_admin():
    """Get the currently logged in admin"""
    admin_id = session.get('admin_id')
    if admin_id:
        return Admin.query.get(admin_id)
    return None


def send_issue_email(issue, user):
    """Send email notification for new issue report"""
    try:
        from flask_mail import Message
        from app import mail

        msg = Message(
            subject=f'New Local Issue Reported: {issue.title}',
            recipients=[current_app.config['MAIL_RECEIVER']],
            sender=current_app.config.get('MAIL_DEFAULT_SENDER') or 'noreply@wardpulse.ai',
            html=f'''
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="background: linear-gradient(135deg, #6366f1, #06b6d4); padding: 20px; border-radius: 12px 12px 0 0;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">🏛️ WardPulse AI</h1>
                    <p style="color: rgba(255,255,255,0.9); margin: 5px 0 0;">New Local Issue Reported</p>
                </div>
                <div style="background: #f8fafc; padding: 24px; border: 1px solid #e2e8f0; border-radius: 0 0 12px 12px;">
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr><td style="padding: 8px 0; font-weight: bold; color: #475569;">Reporter:</td><td style="padding: 8px 0; color: #1e293b;">{user.full_name}</td></tr>
                        <tr><td style="padding: 8px 0; font-weight: bold; color: #475569;">Issue Title:</td><td style="padding: 8px 0; color: #1e293b;">{issue.title}</td></tr>
                        <tr><td style="padding: 8px 0; font-weight: bold; color: #475569;">Category:</td><td style="padding: 8px 0; color: #1e293b;">{issue.category}</td></tr>
                        <tr><td style="padding: 8px 0; font-weight: bold; color: #475569;">Ward Number:</td><td style="padding: 8px 0; color: #1e293b;">{issue.ward}</td></tr>
                        <tr><td style="padding: 8px 0; font-weight: bold; color: #475569;">Location:</td><td style="padding: 8px 0; color: #1e293b;">{issue.address or 'GPS Coordinates'}</td></tr>
                        <tr><td style="padding: 8px 0; font-weight: bold; color: #475569;">Priority:</td><td style="padding: 8px 0; color: #1e293b;">{issue.priority.upper()}</td></tr>
                        <tr><td style="padding: 8px 0; font-weight: bold; color: #475569;">Date:</td><td style="padding: 8px 0; color: #1e293b;">{issue.created_at.strftime('%Y-%m-%d %H:%M')}</td></tr>
                    </table>
                    <div style="margin-top: 16px; padding: 16px; background: white; border-radius: 8px; border: 1px solid #e2e8f0;">
                        <p style="font-weight: bold; color: #475569; margin: 0 0 8px;">Description:</p>
                        <p style="color: #1e293b; margin: 0; line-height: 1.6;">{issue.description}</p>
                    </div>
                </div>
            </div>
            '''
        )

        # Attach image if exists
        if issue.image_path:
            image_full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], issue.image_path)
            if os.path.exists(image_full_path):
                with open(image_full_path, 'rb') as img:
                    msg.attach(issue.image_path, 'image/jpeg', img.read())

        mail.send(msg)
        return True
    except Exception as e:
        current_app.logger.error(f'Email sending failed: {str(e)}')
        return False


def create_notification(user_id, message, notif_type='info'):
    """Create a notification for a user"""
    notif = Notification(user_id=user_id, message=message, type=notif_type)
    db.session.add(notif)
    db.session.commit()


# =============================================
# MAIN ROUTES (Public Pages)
# =============================================
@main_bp.route('/')
def landing():
    """Landing page with hero, stats, features, trending issues"""
    total_issues = Issue.query.count()
    total_users = User.query.count()
    resolved = Issue.query.filter_by(status='resolved').count()
    trending = Issue.query.order_by(Issue.votes_count.desc()).limit(6).all()
    return render_template('landing.html',
                           total_issues=total_issues,
                           total_users=total_users,
                           resolved=resolved,
                           trending=trending)


@main_bp.route('/issues')
def issue_board():
    """Public issue board with search, filter, sort"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    ward = request.args.get('ward', '')
    sort = request.args.get('sort', 'newest')

    query = Issue.query

    # Apply search filters
    if search:
        query = query.filter(
            db.or_(
                Issue.title.ilike(f'%{search}%'),
                Issue.description.ilike(f'%{search}%'),
                Issue.address.ilike(f'%{search}%')
            )
        )
    if category:
        query = query.filter_by(category=category)
    if status:
        query = query.filter_by(status=status)
    if ward:
        query = query.filter_by(ward=ward)

    # Apply sorting
    if sort == 'oldest':
        query = query.order_by(Issue.created_at.asc())
    elif sort == 'votes':
        query = query.order_by(Issue.votes_count.desc())
    elif sort == 'critical':
        query = query.order_by(
            db.case(
                (Issue.priority == 'critical', 1),
                (Issue.priority == 'high', 2),
                (Issue.priority == 'medium', 3),
                (Issue.priority == 'low', 4),
            )
        )
    else:  # newest
        query = query.order_by(Issue.created_at.desc())

    issues = query.paginate(page=page, per_page=12, error_out=False)
    wards = db.session.query(Issue.ward).distinct().order_by(Issue.ward).all()
    ward_list = [w[0] for w in wards]

    return render_template('issues/board.html',
                           issues=issues,
                           search=search,
                           category=category,
                           status=status,
                           ward=ward,
                           sort=sort,
                           ward_list=ward_list)


@main_bp.route('/issues/<int:issue_id>')
def issue_detail(issue_id):
    """Single issue detail page"""
    issue = Issue.query.get_or_404(issue_id)
    has_voted = False
    if current_user.is_authenticated and isinstance(current_user, User):
        has_voted = issue.has_voted(current_user.id)
    return render_template('issues/detail.html', issue=issue, has_voted=has_voted)


@main_bp.route('/map')
def map_view():
    """Full map view of all issues"""
    issues = Issue.query.filter(
        Issue.latitude.isnot(None),
        Issue.longitude.isnot(None)
    ).all()
    issues_data = [i.to_dict() for i in issues]
    return render_template('issues/map_view.html', issues_data=json.dumps(issues_data))


@main_bp.route('/uploads/<filename>')
def uploaded_file(filename):
    """Serve uploaded files"""
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)


# =============================================
# AUTH ROUTES
# =============================================
@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    """User login page"""
    if current_user.is_authenticated and isinstance(current_user, User):
        return redirect(url_for('user.dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = request.form.get('remember', False)

        user = User.query.filter_by(email=email).first()

        if user and user.check_password(password):
            if not user.is_active_user:
                flash('Your account has been disabled. Contact admin.', 'error')
                return redirect(url_for('auth.login'))

            login_user(user, remember=bool(remember))
            user.last_login = datetime.utcnow()
            db.session.commit()
            flash(f'Welcome back, {user.full_name}!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('user.dashboard'))
        else:
            flash('Invalid email or password.', 'error')

    return render_template('auth/login.html')


@auth_bp.route('/register', methods=['GET', 'POST'])
def register():
    """User registration page"""
    if current_user.is_authenticated:
        return redirect(url_for('user.dashboard'))

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip().lower()
        phone = request.form.get('phone', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        # Validation
        errors = []
        if not full_name or len(full_name) < 2:
            errors.append('Full name is required.')
        if not email or '@' not in email:
            errors.append('Valid email is required.')
        if not phone or len(phone) < 10:
            errors.append('Valid phone number is required.')
        if len(password) < 6:
            errors.append('Password must be at least 6 characters.')
        if password != confirm_password:
            errors.append('Passwords do not match.')
        if User.query.filter_by(email=email).first():
            errors.append('Email already registered.')

        if errors:
            for error in errors:
                flash(error, 'error')
            return render_template('auth/register.html')

        # Create user
        user = User(full_name=full_name, email=email, phone=phone)
        user.set_password(password)

        # Handle profile photo
        if 'profile_photo' in request.files:
            file = request.files['profile_photo']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"profile_{email.split('@')[0]}_{int(datetime.utcnow().timestamp())}.{file.filename.rsplit('.', 1)[1].lower()}")
                file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                user.profile_photo = filename

        db.session.add(user)
        db.session.commit()
        flash('Account created successfully! Please log in.', 'success')
        return redirect(url_for('auth.login'))

    return render_template('auth/register.html')


@auth_bp.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if session.get('admin_id'):
        return redirect(url_for('admin.dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')

        admin = Admin.query.filter_by(email=email).first()

        if admin and admin.check_password(password):
            if not admin.is_active_admin:
                flash('Your admin account has been disabled.', 'error')
                return redirect(url_for('auth.admin_login'))

            login_user(admin)  # Sign in with Flask-Login
            session['admin_id'] = admin.id
            session['admin_name'] = admin.full_name
            session['admin_role'] = admin.role
            admin.last_login = datetime.utcnow()
            db.session.commit()
            flash(f'Welcome, {admin.full_name}!', 'success')
            return redirect(url_for('admin.dashboard'))
        else:
            flash('Invalid admin credentials.', 'error')

    return render_template('auth/admin_login.html')


@auth_bp.route('/logout')
def logout():
    """Logout user or admin"""
    if session.get('admin_id'):
        session.pop('admin_id', None)
        session.pop('admin_name', None)
        session.pop('admin_role', None)
        logout_user()  # Sign out of Flask-Login
        flash('Admin logged out successfully.', 'success')
        return redirect(url_for('auth.admin_login'))
    logout_user()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('main.landing'))



@auth_bp.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    """Forgot password page"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user = User.query.filter_by(email=email).first()
        # Always show success message to prevent email enumeration
        flash('If an account exists with that email, password reset instructions have been sent.', 'info')
        return redirect(url_for('auth.login'))
    return render_template('auth/forgot_password.html')


# =============================================
# USER ROUTES (Login Required)
# =============================================
@user_bp.route('/dashboard')
@login_required
def dashboard():
    """User dashboard with stats and recent activity"""
    my_issues = Issue.query.filter_by(user_id=current_user.id).order_by(Issue.created_at.desc()).limit(5).all()
    total_my_issues = Issue.query.filter_by(user_id=current_user.id).count()
    total_votes = current_user.get_total_votes_received()
    pending = Issue.query.filter_by(user_id=current_user.id, status='pending').count()
    resolved = Issue.query.filter_by(user_id=current_user.id, status='resolved').count()
    notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).order_by(Notification.created_at.desc()).limit(10).all()
    trending = Issue.query.order_by(Issue.votes_count.desc()).limit(5).all()

    return render_template('user/dashboard.html',
                           my_issues=my_issues,
                           total_my_issues=total_my_issues,
                           total_votes=total_votes,
                           pending=pending,
                           resolved=resolved,
                           notifications=notifications,
                           trending=trending)


@user_bp.route('/report', methods=['GET', 'POST'])
@login_required
def report_issue():
    """Report a new issue"""
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        category = request.form.get('category', '')
        ward = request.form.get('ward', '').strip()
        address = request.form.get('address', '').strip()
        latitude = request.form.get('latitude', type=float)
        longitude = request.form.get('longitude', type=float)
        priority = request.form.get('priority', 'medium')

        # Validation
        errors = []
        if not title or len(title) < 5:
            errors.append('Issue title must be at least 5 characters.')
        if not description or len(description) < 10:
            errors.append('Description must be at least 10 characters.')
        if category not in Issue.CATEGORIES:
            errors.append('Invalid category.')
        if not ward:
            errors.append('Ward number is required.')
        if priority not in Issue.PRIORITIES:
            priority = 'medium'

        if errors:
            for error in errors:
                flash(error, 'error')
            return render_template('user/report_issue.html')

        # Handle image upload
        image_path = None
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename and allowed_file(file.filename):
                ext = file.filename.rsplit('.', 1)[1].lower()
                filename = secure_filename(f"issue_{current_user.id}_{int(datetime.utcnow().timestamp())}.{ext}")
                file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                image_path = filename

        # Create issue
        issue = Issue(
            title=title,
            description=description,
            category=category,
            ward=ward,
            address=address,
            latitude=latitude,
            longitude=longitude,
            image_path=image_path,
            priority=priority,
            user_id=current_user.id
        )

        db.session.add(issue)
        db.session.commit()

        # Send email notification
        email_sent = send_issue_email(issue, current_user)

        # Create notification for user
        create_notification(
            current_user.id,
            f'Your issue "{title}" has been submitted successfully!',
            'success'
        )

        if email_sent:
            flash('Issue reported successfully! Email notification sent.', 'success')
        else:
            flash('Issue reported successfully! (Email notification could not be sent)', 'success')

        return redirect(url_for('user.my_reports'))

    return render_template('user/report_issue.html')


@user_bp.route('/my-reports')
@login_required
def my_reports():
    """View user's submitted issues"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')

    query = Issue.query.filter_by(user_id=current_user.id)
    if status_filter:
        query = query.filter_by(status=status_filter)

    issues = query.order_by(Issue.created_at.desc()).paginate(page=page, per_page=10, error_out=False)
    return render_template('user/my_reports.html', issues=issues, status_filter=status_filter)


@user_bp.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """User profile page"""
    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'update_profile':
            current_user.full_name = request.form.get('full_name', current_user.full_name).strip()
            current_user.phone = request.form.get('phone', current_user.phone).strip()

            # Handle profile photo update
            if 'profile_photo' in request.files:
                file = request.files['profile_photo']
                if file and file.filename and allowed_file(file.filename):
                    ext = file.filename.rsplit('.', 1)[1].lower()
                    filename = secure_filename(f"profile_{current_user.id}_{int(datetime.utcnow().timestamp())}.{ext}")
                    file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
                    # Delete old profile photo if exists
                    if current_user.profile_photo:
                        old_path = os.path.join(current_app.config['UPLOAD_FOLDER'], current_user.profile_photo)
                        if os.path.exists(old_path):
                            os.remove(old_path)
                    current_user.profile_photo = filename

            db.session.commit()
            flash('Profile updated successfully!', 'success')

        elif action == 'change_password':
            current_password = request.form.get('current_password', '')
            new_password = request.form.get('new_password', '')
            confirm_password = request.form.get('confirm_new_password', '')

            if not current_user.check_password(current_password):
                flash('Current password is incorrect.', 'error')
            elif len(new_password) < 6:
                flash('New password must be at least 6 characters.', 'error')
            elif new_password != confirm_password:
                flash('New passwords do not match.', 'error')
            else:
                current_user.set_password(new_password)
                db.session.commit()
                flash('Password changed successfully!', 'success')

        return redirect(url_for('user.profile'))

    total_issues = Issue.query.filter_by(user_id=current_user.id).count()
    total_votes = current_user.get_total_votes_received()
    recent_issues = Issue.query.filter_by(user_id=current_user.id).order_by(Issue.created_at.desc()).limit(10).all()

    return render_template('user/profile.html',
                           total_issues=total_issues,
                           total_votes=total_votes,
                           recent_issues=recent_issues)


# =============================================
# ADMIN ROUTES
# =============================================
@admin_bp.route('/dashboard')
@admin_required
def dashboard():
    """Admin dashboard with statistics and charts"""
    admin = get_current_admin()
    total_users = User.query.count()
    total_admins = Admin.query.count()
    total_issues = Issue.query.count()
    pending = Issue.query.filter_by(status='pending').count()
    verified = Issue.query.filter_by(status='verified').count()
    assigned = Issue.query.filter_by(status='assigned').count()
    resolved = Issue.query.filter_by(status='resolved').count()
    rejected = Issue.query.filter_by(status='rejected').count()
    critical = Issue.query.filter_by(priority='critical').count()

    # Category distribution for charts
    categories = db.session.query(
        Issue.category, db.func.count(Issue.id)
    ).group_by(Issue.category).all()
    categories = [[c[0], c[1]] for c in categories]

    # Recent complaints
    recent_issues = Issue.query.order_by(Issue.created_at.desc()).limit(10).all()

    # Recent users
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()

    # Ward-wise stats
    ward_stats = db.session.query(
        Issue.ward, db.func.count(Issue.id)
    ).group_by(Issue.ward).order_by(db.func.count(Issue.id).desc()).limit(10).all()

    return render_template('admin/dashboard.html',
                           admin=admin,
                           total_users=total_users,
                           total_admins=total_admins,
                           total_issues=total_issues,
                           pending=pending,
                           verified=verified,
                           assigned=assigned,
                           resolved=resolved,
                           rejected=rejected,
                           critical=critical,
                           categories=categories,
                           recent_issues=recent_issues,
                           recent_users=recent_users,
                           ward_stats=ward_stats)


@admin_bp.route('/complaints')
@admin_required
def complaints():
    """Admin complaint management page"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    ward = request.args.get('ward', '')
    priority = request.args.get('priority', '')

    query = Issue.query

    if search:
        query = query.filter(
            db.or_(
                Issue.title.ilike(f'%{search}%'),
                Issue.description.ilike(f'%{search}%'),
                Issue.ward.ilike(f'%{search}%')
            )
        )
    if category:
        query = query.filter_by(category=category)
    if status:
        query = query.filter_by(status=status)
    if ward:
        query = query.filter_by(ward=ward)
    if priority:
        query = query.filter_by(priority=priority)

    issues = query.order_by(Issue.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    wards = db.session.query(Issue.ward).distinct().order_by(Issue.ward).all()
    ward_list = [w[0] for w in wards]

    return render_template('admin/complaints.html',
                           issues=issues,
                           search=search,
                           category=category,
                           status=status,
                           ward=ward,
                           priority=priority,
                           ward_list=ward_list)


@admin_bp.route('/complaints/<int:issue_id>/status', methods=['POST'])
@admin_required
def update_complaint_status(issue_id):
    """Update complaint status"""
    issue = Issue.query.get_or_404(issue_id)
    new_status = request.form.get('status', '')

    if new_status in Issue.STATUSES:
        old_status = issue.status
        issue.status = new_status
        issue.updated_at = datetime.utcnow()
        db.session.commit()

        # Notify the issue reporter
        create_notification(
            issue.user_id,
            f'Your issue "{issue.title}" status changed from {old_status} to {new_status}.',
            'info'
        )

        flash(f'Issue status updated to {new_status}.', 'success')
    else:
        flash('Invalid status.', 'error')

    return redirect(url_for('admin.complaints'))


@admin_bp.route('/complaints/<int:issue_id>/delete', methods=['POST'])
@admin_required
def delete_complaint(issue_id):
    """Delete a complaint"""
    issue = Issue.query.get_or_404(issue_id)

    # Delete associated image
    if issue.image_path:
        image_full = os.path.join(current_app.config['UPLOAD_FOLDER'], issue.image_path)
        if os.path.exists(image_full):
            os.remove(image_full)

    db.session.delete(issue)
    db.session.commit()
    flash('Complaint deleted successfully.', 'success')
    return redirect(url_for('admin.complaints'))


@admin_bp.route('/users')
@admin_required
def users():
    """User management page"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')

    query = User.query
    if search:
        query = query.filter(
            db.or_(
                User.full_name.ilike(f'%{search}%'),
                User.email.ilike(f'%{search}%'),
                User.phone.ilike(f'%{search}%')
            )
        )

    users_list = query.order_by(User.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/users.html', users=users_list, search=search)


@admin_bp.route('/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def toggle_user(user_id):
    """Enable/disable user account"""
    user = User.query.get_or_404(user_id)
    user.is_active_user = not user.is_active_user
    db.session.commit()
    status = 'enabled' if user.is_active_user else 'disabled'
    flash(f'User {user.full_name} has been {status}.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    """Delete a user account"""
    user = User.query.get_or_404(user_id)
    # Delete profile photo
    if user.profile_photo:
        photo_path = os.path.join(current_app.config['UPLOAD_FOLDER'], user.profile_photo)
        if os.path.exists(photo_path):
            os.remove(photo_path)
    db.session.delete(user)
    db.session.commit()
    flash(f'User {user.full_name} has been deleted.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/admins')
@super_admin_required
def admins():
    """Admin management page (Super Admin only)"""
    admin_list = Admin.query.order_by(Admin.created_at.desc()).all()
    return render_template('admin/admins.html', admins=admin_list)


@admin_bp.route('/admins/create', methods=['POST'])
@super_admin_required
def create_admin():
    """Create a new admin account"""
    full_name = request.form.get('full_name', '').strip()
    email = request.form.get('email', '').strip().lower()
    password = request.form.get('password', '')
    role = request.form.get('role', 'moderator')

    if not full_name or not email or not password:
        flash('All fields are required.', 'error')
        return redirect(url_for('admin.admins'))

    if Admin.query.filter_by(email=email).first():
        flash('Admin with this email already exists.', 'error')
        return redirect(url_for('admin.admins'))

    if role not in ['super_admin', 'ward_officer', 'moderator']:
        role = 'moderator'

    admin = Admin(full_name=full_name, email=email, role=role)
    admin.set_password(password)
    db.session.add(admin)
    db.session.commit()
    flash(f'Admin {full_name} created successfully!', 'success')
    return redirect(url_for('admin.admins'))


@admin_bp.route('/admins/<int:admin_id>/delete', methods=['POST'])
@super_admin_required
def delete_admin(admin_id):
    """Delete an admin account"""
    admin = Admin.query.get_or_404(admin_id)
    if admin.id == session.get('admin_id'):
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('admin.admins'))
    db.session.delete(admin)
    db.session.commit()
    flash(f'Admin {admin.full_name} has been deleted.', 'success')
    return redirect(url_for('admin.admins'))


@admin_bp.route('/admins/<int:admin_id>/toggle', methods=['POST'])
@super_admin_required
def toggle_admin(admin_id):
    """Enable/disable admin account"""
    admin = Admin.query.get_or_404(admin_id)
    if admin.id == session.get('admin_id'):
        flash('You cannot disable your own account.', 'error')
        return redirect(url_for('admin.admins'))
    admin.is_active_admin = not admin.is_active_admin
    db.session.commit()
    status = 'enabled' if admin.is_active_admin else 'disabled'
    flash(f'Admin {admin.full_name} has been {status}.', 'success')
    return redirect(url_for('admin.admins'))


@admin_bp.route('/analytics')
@admin_required
def analytics():
    """AI analytics page"""
    admin = get_current_admin()

    # Ward-wise leaderboard
    ward_stats = db.session.query(
        Issue.ward,
        db.func.count(Issue.id).label('total'),
        db.func.sum(db.case((Issue.status == 'resolved', 1), else_=0)).label('resolved')
    ).group_by(Issue.ward).order_by(db.func.count(Issue.id).desc()).all()

    # Monthly trend
    monthly = db.session.query(
        db.func.strftime('%Y-%m', Issue.created_at).label('month'),
        db.func.count(Issue.id)
    ).group_by('month').order_by('month').all()
    monthly = [[m[0], m[1]] for m in monthly]

    return render_template('admin/analytics.html',
                           admin=admin,
                           ward_stats=ward_stats,
                           monthly=monthly)


@admin_bp.route('/export/excel')
@admin_required
def export_excel():
    """Export complaints to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'Complaints'

        # Header styling
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')

        headers = ['ID', 'Title', 'Category', 'Ward', 'Address', 'Priority',
                    'Status', 'Votes', 'Reporter', 'Date']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data
        issues = Issue.query.order_by(Issue.created_at.desc()).all()
        for row, issue in enumerate(issues, 2):
            ws.cell(row=row, column=1, value=issue.id)
            ws.cell(row=row, column=2, value=issue.title)
            ws.cell(row=row, column=3, value=issue.category)
            ws.cell(row=row, column=4, value=issue.ward)
            ws.cell(row=row, column=5, value=issue.address or '')
            ws.cell(row=row, column=6, value=issue.priority)
            ws.cell(row=row, column=7, value=issue.status)
            ws.cell(row=row, column=8, value=issue.votes_count)
            ws.cell(row=row, column=9, value=issue.reporter.full_name if issue.reporter else 'Unknown')
            ws.cell(row=row, column=10, value=issue.created_at.strftime('%Y-%m-%d %H:%M'))

        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'wardpulse_complaints_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('admin.complaints'))


@admin_bp.route('/export/pdf')
@admin_required
def export_pdf():
    """Export complaints to PDF file"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph('WardPulse AI - Complaints Report', styles['Title']))
        elements.append(Paragraph(f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        elements.append(Spacer(1, 20))

        # Table data
        headers = ['ID', 'Title', 'Category', 'Ward', 'Priority', 'Status', 'Votes', 'Date']
        data = [headers]

        issues = Issue.query.order_by(Issue.created_at.desc()).all()
        for issue in issues:
            data.append([
                str(issue.id),
                issue.title[:40],
                issue.category,
                issue.ward,
                issue.priority,
                issue.status,
                str(issue.votes_count),
                issue.created_at.strftime('%Y-%m-%d')
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(table)
        doc.build(elements)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'wardpulse_complaints_{datetime.utcnow().strftime("%Y%m%d")}.pdf'
        )
    except Exception as e:
        flash(f'PDF export failed: {str(e)}', 'error')
        return redirect(url_for('admin.complaints'))


@admin_bp.route('/download-image/<filename>')
@admin_required
def download_image(filename):
    """Download uploaded image"""
    return send_from_directory(
        current_app.config['UPLOAD_FOLDER'],
        filename,
        as_attachment=True
    )


# =============================================
# API ROUTES (AJAX Endpoints)
# =============================================
@api_bp.route('/vote/<int:issue_id>', methods=['POST'])
@login_required
def vote(issue_id):
    """Vote/unvote on an issue"""
    issue = Issue.query.get_or_404(issue_id)

    if not isinstance(current_user, User):
        return jsonify({'error': 'Only users can vote'}), 403

    existing_vote = Vote.query.filter_by(issue_id=issue_id, user_id=current_user.id).first()

    if existing_vote:
        # Remove vote
        db.session.delete(existing_vote)
        issue.votes_count = max(0, issue.votes_count - 1)
        db.session.commit()
        return jsonify({'voted': False, 'votes_count': issue.votes_count})
    else:
        # Add vote
        vote = Vote(issue_id=issue_id, user_id=current_user.id)
        db.session.add(vote)
        issue.votes_count += 1
        db.session.commit()

        create_notification(
            issue.user_id,
            f'{current_user.full_name} upvoted your issue: "{issue.title}"',
            'success'
        )

        return jsonify({'voted': True, 'votes_count': issue.votes_count})


@api_bp.route('/issues')
def get_issues():
    """JSON API for issues (used by map and AJAX)"""
    issues = Issue.query.filter(
        Issue.latitude.isnot(None),
        Issue.longitude.isnot(None)
    ).all()
    return jsonify([i.to_dict() for i in issues])


@api_bp.route('/notifications')
@login_required
def get_notifications():
    """Get user's unread notifications"""
    notifications = Notification.query.filter_by(
        user_id=current_user.id, is_read=False
    ).order_by(Notification.created_at.desc()).limit(20).all()
    return jsonify([n.to_dict() for n in notifications])


@api_bp.route('/notifications/read', methods=['POST'])
@login_required
def mark_notifications_read():
    """Mark all notifications as read"""
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'success': True})


@api_bp.route('/chatbot', methods=['POST'])
def chatbot():
    """AI Chatbot endpoint using Gemini API"""
    message = request.json.get('message', '').strip()
    if not message:
        return jsonify({'response': 'Please enter a message.'})

    api_key = current_app.config.get('GEMINI_API_KEY', '')

    if not api_key:
        # Fallback responses when no API key is configured
        return jsonify({'response': get_fallback_response(message)})

    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')

        # System context for the chatbot
        system_prompt = """You are WardPulse AI Assistant, a helpful chatbot for a civic complaint portal.
        You help citizens with:
        1. How to report issues (go to Dashboard > Report Issue, fill form with title, description, category, ward, location, image)
        2. Government schemes for urban development
        3. Emergency contacts (Police: 100, Fire: 101, Ambulance: 102, Municipal: 1800-XXX-XXXX)
        4. How voting works (login, go to Issue Board, click upvote button - one vote per issue)
        5. Understanding complaint status (Pending → Verified → Assigned → Resolved)
        6. Categories: Road Damage, Garbage, Street Light, Drainage, Water Leakage, Electricity, Public Safety, Tree Fallen
        Keep responses concise, friendly, and helpful. Use emojis sparingly."""

        response = model.generate_content(f"{system_prompt}\n\nUser: {message}")
        return jsonify({'response': response.text})
    except Exception as e:
        current_app.logger.error(f'Chatbot error: {str(e)}')
        return jsonify({'response': get_fallback_response(message)})


@api_bp.route('/ai/detect-category', methods=['POST'])
def detect_category():
    """AI-based category detection from description"""
    description = request.json.get('description', '').strip()
    if not description:
        return jsonify({'category': 'Others', 'priority': 'medium'})

    api_key = current_app.config.get('GEMINI_API_KEY', '')

    if not api_key:
        return jsonify(detect_category_fallback(description))

    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')

        prompt = f"""Analyze this civic complaint and return ONLY a JSON object with two fields:
        1. "category": one of [Road Damage, Garbage, Street Light, Drainage, Water Leakage, Electricity, Public Safety, Tree Fallen, Others]
        2. "priority": one of [critical, high, medium, low]

        Complaint: {description}

        Return ONLY the JSON object, no other text."""

        response = model.generate_content(prompt)
        text = response.text.strip()
        # Try to parse JSON from response
        if '{' in text:
            text = text[text.index('{'):text.rindex('}') + 1]
            result = json.loads(text)
            return jsonify(result)
    except Exception as e:
        current_app.logger.error(f'AI category detection error: {str(e)}')

    return jsonify(detect_category_fallback(description))


@api_bp.route('/ai/check-duplicate', methods=['POST'])
def check_duplicate():
    """Check for duplicate/similar issues"""
    title = request.json.get('title', '').strip().lower()
    description = request.json.get('description', '').strip().lower()

    if not title:
        return jsonify({'duplicate': False})

    # Simple keyword-based duplicate check
    similar = Issue.query.filter(
        db.or_(
            Issue.title.ilike(f'%{title}%'),
            Issue.title.ilike(f'%{" ".join(title.split()[:3])}%')
        )
    ).limit(3).all()

    if similar:
        return jsonify({
            'duplicate': True,
            'similar_issues': [{'id': i.id, 'title': i.title, 'votes': i.votes_count, 'status': i.status} for i in similar]
        })

    return jsonify({'duplicate': False})


@api_bp.route('/ai/summary', methods=['POST'])
@admin_required
def ai_summary():
    """Generate AI summary for admin dashboard"""
    api_key = current_app.config.get('GEMINI_API_KEY', '')

    total = Issue.query.count()
    pending = Issue.query.filter_by(status='pending').count()
    resolved = Issue.query.filter_by(status='resolved').count()
    critical = Issue.query.filter_by(priority='critical').count()

    # Get top categories
    top_cats = db.session.query(
        Issue.category, db.func.count(Issue.id)
    ).group_by(Issue.category).order_by(db.func.count(Issue.id).desc()).limit(3).all()

    if not api_key:
        summary = f"""📊 Daily Summary:
• Total Issues: {total}
• Pending: {pending} | Resolved: {resolved}
• Critical Issues: {critical}
• Top Categories: {', '.join([f'{c[0]} ({c[1]})' for c in top_cats])}
• Resolution Rate: {round(resolved/total*100, 1) if total > 0 else 0}%

💡 Recommendation: Focus on the {pending} pending issues, especially {critical} critical ones that need immediate attention."""
        return jsonify({'summary': summary})

    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')

        prompt = f"""Generate a brief admin dashboard summary for a civic complaints portal:
        - Total Issues: {total}
        - Pending: {pending}, Resolved: {resolved}
        - Critical Issues: {critical}
        - Top Categories: {', '.join([f'{c[0]} ({c[1]})' for c in top_cats])}

        Include key insights, trends, and recommendations. Keep it concise (3-4 bullet points).
        Use emojis for visual appeal."""

        response = model.generate_content(prompt)
        return jsonify({'summary': response.text})
    except Exception as e:
        return jsonify({'summary': f'AI summary unavailable: {str(e)}'})


@api_bp.route('/qrcode/<int:issue_id>')
def generate_qr(issue_id):
    """Generate QR code for an issue"""
    try:
        import qrcode
        issue = Issue.query.get_or_404(issue_id)
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(f'{request.host_url}issues/{issue_id}')
        qr.make(fit=True)
        img = qr.make_image(fill_color='#6366f1', back_color='white')

        output = io.BytesIO()
        img.save(output, format='PNG')
        output.seek(0)

        return send_file(output, mimetype='image/png')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@api_bp.route('/stats')
def get_stats():
    """Get live statistics"""
    return jsonify({
        'total_issues': Issue.query.count(),
        'total_users': User.query.count(),
        'resolved': Issue.query.filter_by(status='resolved').count(),
        'pending': Issue.query.filter_by(status='pending').count(),
        'critical': Issue.query.filter_by(priority='critical').count()
    })


# =============================================
# Fallback Functions (when no AI API key)
# =============================================
def get_fallback_response(message):
    """Provide helpful responses without AI API"""
    message_lower = message.lower()

    if any(w in message_lower for w in ['report', 'complaint', 'submit', 'issue']):
        return "📝 To report an issue:\n1. Log in to your account\n2. Go to Dashboard → Report Issue\n3. Fill in the title, description, category, ward number\n4. Add your location (GPS or manual)\n5. Upload a photo of the issue\n6. Select priority and submit!\n\nYour complaint will be tracked and you'll receive updates."

    elif any(w in message_lower for w in ['vote', 'upvote', 'like']):
        return "👍 How voting works:\n1. Browse the Issue Board\n2. Find issues you want to support\n3. Click the upvote button (▲)\n4. Each user can vote once per issue\n5. Higher voted issues get more attention from authorities!"

    elif any(w in message_lower for w in ['status', 'track', 'progress']):
        return "📊 Complaint Status Flow:\n• **Pending** → Complaint received\n• **Verified** → Issue confirmed by officer\n• **Assigned** → Assigned to relevant department\n• **Resolved** → Issue fixed\n• **Rejected** → Invalid or duplicate complaint\n\nCheck your dashboard for real-time updates!"

    elif any(w in message_lower for w in ['emergency', 'urgent', 'help', 'contact']):
        return "🚨 Emergency Contacts:\n• Police: 100\n• Fire Brigade: 101\n• Ambulance: 102\n• Women Helpline: 1091\n• Municipal Corporation: 1800-XXX-XXXX\n• Disaster Management: 108\n\nFor non-emergency civic issues, please use the Report Issue feature."

    elif any(w in message_lower for w in ['category', 'type', 'categories']):
        return "📋 Issue Categories:\n• 🛣️ Road Damage\n• 🗑️ Garbage\n• 💡 Street Light\n• 🌊 Drainage\n• 💧 Water Leakage\n• ⚡ Electricity\n• 🛡️ Public Safety\n• 🌳 Tree Fallen\n• 📌 Others"

    elif any(w in message_lower for w in ['hello', 'hi', 'hey', 'namaste']):
        return "👋 Hello! I'm WardPulse AI Assistant. I can help you with:\n• Reporting civic issues\n• Understanding complaint status\n• Emergency contacts\n• Voting on issues\n• Government schemes\n\nWhat would you like to know?"

    elif any(w in message_lower for w in ['scheme', 'government', 'yojana']):
        return "🏛️ Key Government Schemes:\n• **Smart City Mission** - Urban development\n• **Swachh Bharat** - Cleanliness & sanitation\n• **AMRUT** - Urban infrastructure\n• **PMAY** - Housing for all\n• **Digital India** - E-governance\n\nVisit india.gov.in for more details."

    else:
        return "🤖 I can help you with:\n1. 📝 How to report issues\n2. 👍 How voting works\n3. 📊 Understanding complaint status\n4. 🚨 Emergency contacts\n5. 📋 Issue categories\n6. 🏛️ Government schemes\n\nPlease ask me about any of these topics!"


def detect_category_fallback(description):
    """Simple keyword-based category detection"""
    desc_lower = description.lower()

    category_keywords = {
        'Road Damage': ['road', 'pothole', 'crack', 'pavement', 'highway', 'street damage', 'broken road'],
        'Garbage': ['garbage', 'trash', 'waste', 'litter', 'dump', 'rubbish', 'dirty', 'filth'],
        'Street Light': ['light', 'lamp', 'dark', 'bulb', 'street light', 'pole light', 'no light'],
        'Drainage': ['drain', 'sewer', 'flood', 'waterlog', 'clog', 'blocked drain', 'overflow'],
        'Water Leakage': ['water', 'leak', 'pipe', 'burst', 'tap', 'supply', 'water leak'],
        'Electricity': ['electric', 'power', 'wire', 'transformer', 'outage', 'voltage', 'current'],
        'Public Safety': ['safety', 'danger', 'accident', 'crime', 'theft', 'unsafe', 'security'],
        'Tree Fallen': ['tree', 'branch', 'fallen', 'uprooted', 'broken tree']
    }

    for category, keywords in category_keywords.items():
        if any(kw in desc_lower for kw in keywords):
            # Determine priority
            priority = 'medium'
            critical_words = ['emergency', 'urgent', 'danger', 'accident', 'collapse', 'fire', 'flood']
            high_words = ['broken', 'severe', 'major', 'big', 'critical']
            low_words = ['minor', 'small', 'slight', 'little']

            if any(w in desc_lower for w in critical_words):
                priority = 'critical'
            elif any(w in desc_lower for w in high_words):
                priority = 'high'
            elif any(w in desc_lower for w in low_words):
                priority = 'low'

            return {'category': category, 'priority': priority}

    return {'category': 'Others', 'priority': 'medium'}
